#!/usr/bin/env python3

#the data is sent in an average space of 15 seconds, so the basic script runs every 2 minutes, get the last 8 observations calcualte average and send EMail/SMS
# For daily report we can send the average , max and min along with recommended range
# same for weekly report
# Now I have to figure out how to make a snooze button in both SMS and email

import datetime # in the same manner we import requests
import requests
from functools import reduce

#Necessary libraries to be imported for sending emails
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

#Necessary libraries to be imported for sending SMS
from twilio.rest import Client

#read email imports
import imaplib
import email
from email.header import decode_header
from bs4 import BeautifulSoup
import re
import openpyxl

# Email alert
# definition of email IDs
sender_email = 'pradeesh.nit@gmail.com'
receive_email = 'keshav.875@gmail.com'
receiver_email = 'ge84zon@mytum.de'     # Incoming emails from this ID will be read to update notifications
password = 'qczkdohedxmeqizr'           # API password generated specifcally for email functionality

def send_email(result):
    sender_email = 'pradeesh.nit@gmail.com'
    receive_email = 'keshav.875@gmail.com'
    receiver_email = 'ge84zon@mytum.de'  # Incoming emails from this ID will be read to update notifications
    password = 'qczkdohedxmeqizr'  # API password generated specifcally for email functionality
    subject = result
    body_text = result
    # HTML Formatting for a bettter visual appeal
    body_text_html = f"""   
    <html>
      <body>
        <div  style="width: 50%; margin-left: 20%;padding:50px; background-image: linear-gradient(to right,#ffecd2 0%, #fcb69f 100%);">
            <h1>Indoor Air Quality Alert !!! <h1>
            <h2 style="color: red;">{body_text}<h2>
            <h3>Visit <a style="color: green;" href="https://gi3.gis.lrg.tum.de/grafana/d/yqQ2KwX4k/pradeesh-dashboard-voc?orgId=1">Dashboard</a> for more details<h3>
            <p>To change notification preferences, reply   1 - Stop email, 2- Stop SMS, 3 - Stop both, 4 - Resume email, 5 - Resume SMS, 6 - Resume both</p>       
        </div>
      </body>
    </html>
    """
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = receive_email
    message['Subject'] = subject
    message.attach(MIMEText(body_text_html, 'html'))

    smtp_server = 'smtp.gmail.com'
    smtp_port = 587

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message.as_string())
        server.quit()
        print('Email sent successfully!')
    except smtplib.SMTPException as e:
        print('Error sending email:', str(e))

# SMS Alert - for high priority
# +14782862242         # SMS alert will be sent from this fixed number. To receive SMS, prior number verification via OTP is required

def send_sms(result):
    account_sid = 'AC4203cf620566720b4f512a33bb220ea1'
    auth_token = '6515bb8d036489e782faf272163a4795'
    try:
        client = Client(account_sid, auth_token)
        message = client.messages.create(
            from_='+14782862242',
            body=result,
            # to='+4915510576219'
            to='+4917647388347'
        )
        print('SMS sent successfully!')
        print(message.sid)
    except OSError as e:
        print(e)

# Specifying datastreams to be queried and their safe limits
datasstreams = [826,827,828,829, 830, 831, 832, 876, 877]
datastream_names = ['Temperature', 'Humidity', 'Air Pressure', 'VOC Gas','PM1','PM2.5','PM10','AirQuality','AirQualityNumber']
data_limits = [[16, 26], [20, 65], [95, 102.26], [70, 500], [0, 15], [0, 15], [0, 45], [2, 5], [2, 100]]


# Make API Queries to the FROST Server and find the average result over a period of time
def averageResult(id, index):
    URL = f'https://gi3.gis.lrg.tum.de/frost/v1.1/Datastreams({id})/Observations?$select=result,phenomenonTime&$orderby=phenomenonTime%20desc&$top=10'  #$select=result,phenomenonTime&
    result = requests.get(url = URL)
    data = result.json()
    # print(data)
    avg = reduce((lambda x, y: x + y), list(map(lambda x: x['result'], data['value'])))/len(data['value'])
    print(f'Datastream ID: {id}, Average value: {avg}', end=" ")
    result = decision(datastream_names[index],avg,data_limits[index][0],data_limits[index][1])
    return result

# Process the Data and Make Decisions
def decision(property, value, low, high):
    result = f'{property} too low !!! ' if value < low else f'{property} too high !!! ' if value > high else ""
    if(result != ""):
        print(f'| x {result}')
    else :
        print("| ✓ ok  ")
    return result

#helper functions to read email
def extract_numbers(text):
	numbers = re.findall(r'\b\d+\b', text)
	return numbers

def get_plain_text_from_html(html_content):
	soup = BeautifulSoup(html_content, 'html.parser')
	plain_text = soup.get_text()
	return plain_text

def get_email_body(message):
    email_body = ""

    for part in message.walk():
        content_type = part.get_content_type()
        content_disposition = str(part.get("Content-Disposition"))

        if "attachment" not in content_disposition:
            try:
                payload = part.get_payload(decode=True)
                if isinstance(payload, bytes):
                    charset = part.get_content_charset()
                    if charset is None:
                        charset = "utf-8"  # Default to utf-8 if charset is not specified
                    content = payload.decode(charset, errors="ignore")

                    if content_type == "text/plain":
                        email_body += content  # Only process plain text content
            except Exception as e:
                print("Error decoding part:", str(e))

    return email_body


def get_body(msg):
	if msg.is_multipart():
		return get_body(msg.get_payload(0))
	else:
		return msg.get_payload(None, True)


def search(key, value, con):
	result, data = con.search(None, key, '"{}"'.format(value))
	return data


def get_emails(result_bytes):
	msgs = []
	for num in result_bytes[0].split():
		typ, data = con.fetch(num, '(RFC822)')
		msgs.append(data)
	return msgs

def update_notifcation_status():
    user = sender_email
    pw = password
    imap_url = 'imap.gmail.com'
    con = imaplib.IMAP4_SSL(imap_url)
    con.login(user, pw)
    con.select('Inbox')
    search_email = receiver_email
    result, data = con.search(None, 'FROM', f'"{search_email}"')
    latest_email_num = data[0].split()[-1]  # Get the latest email number
    typ, msg_data = con.fetch(latest_email_num, '(RFC822)')
    msg = email.message_from_bytes(msg_data[0][1])
    subject, _ = decode_header(msg["Subject"])[0]
    subject = subject.decode("utf-8") if isinstance(subject, bytes) else subject
    email_body = get_email_body(msg)
    extracted_numbers = extract_numbers(email_body)[0]
    # print("Email Body:", email_body)
    # print('email number', extracted_numbers)
    state_changes = {
        "1": "01",  # Stop email
        "2": "10",  # Stop SMS
        "3": "00",  # Stop both
        "4": "10",  # Resume email
        "5": "01",  # Resume SMS
        "6": "11"  # Resume both
    }

    # Get the existing state from the Excel file
    excel_file_path = "notification_states.xlsx"
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    if (extracted_numbers == '2') or (extracted_numbers == '5'):
        ws["B1"] = ws["B1"].value if ws["B1"].value == state_changes.get(extracted_numbers)[1] else \
        state_changes.get(extracted_numbers)[1]
    elif (extracted_numbers == '1') or (extracted_numbers == '4'):
        ws["A1"] = ws["A1"].value if ws["A1"].value == state_changes.get(extracted_numbers)[0] else \
        state_changes.get(extracted_numbers)[0]
    else:
        ws["B1"] = ws["B1"].value if ws["B1"].value == state_changes.get(extracted_numbers)[1] else \
        state_changes.get(extracted_numbers)[1]
        ws["A1"] = ws["A1"].value if ws["A1"].value == state_changes.get(extracted_numbers)[0] else \
        state_changes.get(extracted_numbers)[0]
    wb.save(excel_file_path)

def send_alert(result):
    update_notifcation_status()
    excel_file_path = "notification_states.xlsx"
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    if ws["A1"].value == '1':
        send_email(result)
    else:
        print('Email notificaton snoozed')
    if ws["B1"].value == '1':
        send_sms(result)
    else:
        print('SMS notificaton snoozed')




#Function call
answer = ""
for i in range(len(datasstreams)):
  print(f'Property name: {datastream_names[i]}, ', end=" ")
  temp = averageResult(datasstreams[i],i)
  if(temp != ""):
      answer += temp + " "
if(answer != ""):
    send_alert(answer)  # This functions handles alerting users via SMS and email based on user preferences
